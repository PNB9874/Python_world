import openpyxl

path = "C:\\Users\\P N Babu\\Desktop\\demo.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb["Sheet1"]


rowCount = sheet.max_row
colCount = sheet.max_column

for row in range(1,rowCount+1):
    for col in range(1, colCount+1):
        print(sheet.cell(row,col).value,end='   ')
    print()